import os
import glob
import cv2
import numpy as np
import pandas as pd
from scipy.optimize import curve_fit
from scipy.stats import linregress
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import seaborn as sns


palette = sns.color_palette("bright")

def process_image(image_bg, thresh_factor):
    """
    Processus de traitement d'une image pour détecter les contours 
    dans la densité spectrale.
    
    Args:
        image: np.ndarray, l'image en niveaux de gris
        thresh_factor: float, facteur de seuillage pour déterminer 
        le seuil de binarisation
        
    Returns:
        list, une liste de contours des régions d'intérêt fermées
    """
    img8 = np.clip((image_bg/65535.0) * 255.0, 0, 255).astype(np.uint8)

    threshold = thresh_factor * np.max(img8)
    regions_of_interest = (img8 > threshold).astype(np.uint8)

    kernel = np.ones((5, 5), np.uint8)
    regions_of_interest_closed = cv2.morphologyEx(regions_of_interest,
                                                cv2.MORPH_CLOSE, kernel)

    contours, _ = cv2.findContours(regions_of_interest_closed,
                                cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    filtered_contours = []
    for contour in contours:
        _, _, w, h = cv2.boundingRect(contour)
        if w > 10 and h > 10:
            filtered_contours.append(contour)

    return filtered_contours

def draw_all_contours(image, image_bg, name, factor=0.02):
    """
    Dessine les contours des régions d'intérêt sur l'image de densité spectrale.
    
    Args:
        image: np.ndarray, l'image en niveaux de gris
        name: string,  nom du fichier à sauvegarder
        factor: float, facteur de seuillage pour déterminer le seuil de binarisation
    """
    cnt = process_image(image_bg, factor)
    image = np.clip((image/65535.0) * 255.0, 0, 255).astype(np.uint8)
    colored_spectrum = cv2.cvtColor(image.astype(np.uint8)
                                    , cv2.COLOR_GRAY2BGR)

    for i, contour in enumerate(cnt):
        cv2.drawContours(colored_spectrum, [contour], -1, (0, 0, 255), 2)
        contour_center = contour.mean(axis=0).astype(int)[0]
        cv2.putText(colored_spectrum, str(i+1), tuple(contour_center),
                    cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (255, 255, 255), 2)

    directory = os.path.join("output", "01_all_contours")
    if not os.path.exists(directory):
            os.makedirs(directory)

    cv2.imwrite(os.path.join(directory, name + ".png"), colored_spectrum)

def remove_contours(image_bg, contours_to_remove, factor=0.02):
    """
    Remove specified contours from the list.
    
    Args:
        image: np.ndarray, the grayscale image
        contours_to_remove: list, list of contour indices to remove
        factor: float, thresholding factor to determine the binarization threshold
    """
    bad_cnt = process_image(image_bg, factor)
    good_cnt = [contour for i, contour in enumerate(bad_cnt)
                if i+1 not in contours_to_remove]

    return good_cnt

def draw_good_contours(image, name, contours):
    """
    Draw only good contours on an image and save it.

    Args:
        image (np.ndarray): Input grayscale image.
        name (str): Name of the output image file.
        contours (list): List of contours to draw on the image.
    """
    image = np.clip((image/65535.0) * 255.0, 0, 255).astype(np.uint8)
    colored_spectrum = cv2.cvtColor(image.astype(np.uint8)
                                    , cv2.COLOR_GRAY2BGR)

    for i, contour in enumerate(contours):
        cv2.drawContours(colored_spectrum, [contour], -1, (0, 0, 255), 2)
        contour_center = contour.mean(axis=0).astype(int)[0]
        cv2.putText(colored_spectrum, str(i+1), tuple(contour_center)
                    , cv2.FONT_HERSHEY_COMPLEX_SMALL, 1, (255, 255, 255), 2)

    directory = os.path.join("output", "02_contours_removed")
    os.makedirs(directory, exist_ok=True)
    cv2.imwrite(os.path.join(directory, name + ".png"), colored_spectrum)

def gaussian(x, a, x0, sigma):
    """
    Gaussian function.
    
    Args:
        x (array-like): Input data.
        a (float): Amplitude of the Gaussian curve.
        x0 (float): Mean of the Gaussian curve.
        sigma (float): Standard deviation of the Gaussian curve.
        
    Returns:
        array-like: Values of the Gaussian function evaluated at x.
    """
    return a * np.exp(-(x - x0)**2 / (2 * sigma**2))

def find_contour_centroids(contours, image):
    """
    Find centroids of contours in an image.

    Args:
        contours (list): List of contours.
        image (np.ndarray): Grayscale image.

    Returns:
        tuple: A tuple containing two lists. The first list contains
        tuples of centroid coordinates (x, y) for each contour.
            The second list contains tuples of centroid uncertainties
            (x_uncertainty, y_uncertainty) for each contour.
    """
    centroids = []
    centroids_uncertainty = []
    for i, contour in enumerate(contours):
        x_coords = contour[:, 0, 0]
        y_coords = contour[:, 0, 1]

        min_x_coord = np.min(x_coords)
        max_x_coord = np.max(x_coords)
        min_y_coord = np.min(y_coords)
        max_y_coord = np.max(y_coords)

        pixel_values = image[min_y_coord:max_y_coord, min_x_coord:max_x_coord]

        x_mean_intensity = np.mean(pixel_values, axis=0)
        y_mean_intensity = np.mean(pixel_values, axis=1)

        x_data = np.arange(len(x_mean_intensity))
        y_data = np.arange(len(y_mean_intensity))
        try:
            popt_x, cov_x = curve_fit(gaussian, x_data, x_mean_intensity, 
                                    p0=[np.max(x_mean_intensity), len(x_data) / 2, 10])
            popt_y, cov_y = curve_fit(gaussian, y_data, y_mean_intensity, 
                                    p0=[np.max(y_mean_intensity), len(y_data) / 2, 10])

            x_center = int(popt_x[1])
            x_center_pix = min_x_coord + x_center
            x_center_uncertainty = np.sqrt(np.diag(cov_x)[1])
            y_center = int(popt_y[1])
            y_center_pix = min_y_coord + y_center
            y_center_uncertainty = np.sqrt(np.diag(cov_y)[1])

            centroids.append((x_center_pix, y_center_pix))
            centroids_uncertainty.append((x_center_uncertainty, y_center_uncertainty))

        except RuntimeError:
            x_center = int(len(x_coords) // 2)
            x_center_pix = min_x_coord + x_center
            x_center_uncertainty = np.sqrt(len(x_coords)/2)
            y_center = int(len(y_coords) // 2)
            y_center_pix = min_y_coord + y_center
            y_center_uncertainty = np.sqrt(len(y_coords)/2)

            centroids.append((x_center_pix, y_center_pix))
            centroids_uncertainty.append((x_center_uncertainty, y_center_uncertainty))

            print(f"Error occured during curve fitting for contour {i+1}. Skipping contour.")
            continue

    return centroids, centroids_uncertainty

def draw_points(image, name, centroids, laue_dict, center):
    """
    Draws points at the centroids in the image and saves it as an image 
    file with the specified name.
    
    Parameters:
    image (numpy.ndarray): The image to draw on.
    name (str): The base name of the output image file.
    centroids (list): The list of centroids to draw.
    """
    image = np.clip((image/65535.0) * 255.0, 0, 255).astype(np.uint8)
    colored_spectrum = cv2.cvtColor(image.astype(np.uint8), 
                                    cv2.COLOR_GRAY2BGR)

    h = laue_dict['h']
    k = laue_dict['k']
    l = laue_dict['l']

    centroids.remove(center)

    for i, centroid in enumerate(centroids):
        x, y = map(int, centroid)

        cv2.circle(colored_spectrum, (x, y), 10, (255, 255, 255), -1)

        text = str((h[i], k[i], l[i]))
        text_size = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 1)[0]

        text_x = x - text_size[0] // 2
        text_y = y + 40

        # Ensure text doesn't go outside the image boundaries
        if text_x < 0:
            text_x = 0
        elif text_x + text_size[0] > image.shape[1]:
            text_x = image.shape[1] - text_size[0]

        if text_y < 0:
            text_y = 0
        elif text_y + text_size[1] > image.shape[0]:
            text_y = image.shape[0] - text_size[1]

        cv2.putText(colored_spectrum, text, (text_x, text_y),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1, cv2.LINE_AA)

    out_dir = os.path.join("output", "05_hkl")
    os.makedirs(out_dir, exist_ok=True)
    cv2.imwrite(os.path.join(out_dir, name + ".png"), colored_spectrum)

def find_pairs(image, centroids):
    """
    Trouve les paires de centroïdes symétriques.


    Args:
        image (np.ndarray): L'image.
        centroids (list): Liste des coordonnées des centroïdes à rechercher.

    Returns:
        list: Liste de paires de centroïdes symétriques.
    """
    pos_pairs = []
    centroids2match = centroids.copy()

    for c in centroids2match:
        y, x = c

        center_y = int(image.shape[0] / 2) - 110 <= y <= int(image.shape[0] / 2) + 110
        center_x = int(image.shape[1] / 2) - 110 <= x <= int(image.shape[1] / 2) + 110
        if center_y and center_x:
            center = (y, x)
            centroids2match.remove((y, x))
            break

    for c in centroids2match:
        y, x = c
    
        y_prime = 2 * (center[0] - y) + y
        x_prime = 2 * (center[1] - x) + x

        for c_prime in centroids2match:
            y_prime_centroid, x_prime_centroid = c_prime
            if abs(y_prime_centroid - y_prime) <= 50 and abs(x_prime_centroid
                                                            - x_prime) <= 50:
                pos_pairs.append((c, c_prime))
                centroids2match.remove(c_prime)
                break

    return pos_pairs, center

def draw_pairs(image, name, pairs):
    """
    Draw pairs of points.

    Args:
        image (np.ndarray): Input grayscale image.
        name (str): Name of the output image file.
        pairs (list): List of pairs of points to draw on the image.
    """
    image = np.clip((image/65535.0) * 255.0, 0, 255).astype(np.uint8)
    colored_spectrum = cv2.cvtColor(image.astype(np.uint8),
                                    cv2.COLOR_GRAY2BGR)

    for i, (start_point, end_point) in enumerate(pairs):
        color = tuple(np.random.randint(0, 256, 3).tolist())
        cv2.line(colored_spectrum, start_point, end_point, color,
                1, cv2.LINE_AA)
        for point in [start_point, end_point]:
            cv2.putText(colored_spectrum, str(i + 1), point,
                        cv2.FONT_HERSHEY_SIMPLEX, 1, color, 2)

    out_dir = os.path.join("output", "03_pairs")
    os.makedirs(out_dir, exist_ok=True)
    cv2.imwrite(os.path.join(out_dir, name + ".png"), colored_spectrum)

def laue(name, centroids, uncerts, center):
    """
    Calculates various parameters related to crystallography based on input data.

    Args:
        name (str): The name of the crystal. Should be one of "LiF", "NaCl", or "Si".
        centroids (list): A list of centroid coordinates (x, y) for each point.
        uncerts (list): A list of uncertainty values for each centroid.
        center (tuple): The centroid coordinates (x, y) of the center point.

    Returns:
        dict: A dictionary containing the calculated parameters including:
            - 'xQ': List of x-coordinates after centering.
            - 'xQ_err': List of uncertainties in x-coordinates.
            - 'yQ': List of y-coordinates after centering.
            - 'yQ_err': List of uncertainties in y-coordinates.
            - 'zQ': List of z-coordinates.
            - 'zQ_err': List of uncertainties in z-coordinates.
            - 'fct': List of factors.
            - 'h': List of h-values.
            - 'k': List of k-values.
            - 'l': List of l-values.
            - 'n': List of n-values.
            - 'd_hkl': List of d_hkl values.
            - 'theta': List of theta values.
            - 'wavelength': List of wavelengths.
    """
    # Define crystal lattice constants
    crystal_constants_theo = {"LiF": 403E-12, "NaCl": 564E-12, "Si": 543E-12}
    a0_theo = crystal_constants_theo.get(name, 0)

    crystal_constants_exp = {"LiF": 404.1776E-12, "NaCl": 564.6809E-12, "Si": 0}
    crystal_constants_err = {"LiF": 0.2532E-12, "NaCl": 0.2035E-12, "Si": 0}
    a0_exp = crystal_constants_exp.get(name, 0)
    a0_err = crystal_constants_err.get(name, 0)

    # Constants for calculations
    resolution = 49.5E-6
    L = 2E-2

    # Initialize laue_data dictionary
    laue_data = {
        'xQ': [],
        'xQ_err': [],
        'yQ': [],
        'yQ_err': [],
        'zQ': [],
        'zQ_err': [],
        'fct': [],
        'h': [],
        'k': [],
        'l': [],
        'n': [],
        'd_hkl_theo': [],
        'd_hkl_exp': [],
        'd_hkl_err': [],
        'theta': [],
        'wavelength_theo': [],
        'wavelength_exp': [],
        'wavelength_err': []
    }

    # Calculate center index
    center_index = centroids.index(center)

    # Populate 'xQ', 'xQ_err', 'yQ', 'yQ_err' lists
    for i, centroid in enumerate(centroids):
        if i == center_index:
            continue
        x, y = map(int, centroid)
        xQ = (x - center[1]) * resolution
        xQ_uncert = (uncerts[i][0] + uncerts[center_index][0]) * resolution
        yQ = (y - center[0]) * resolution
        yQ_uncert = (uncerts[i][1] + uncerts[center_index][1]) * resolution

        laue_data['xQ'].append(xQ)
        laue_data['xQ_err'].append(xQ_uncert)
        laue_data['yQ'].append(yQ)
        laue_data['yQ_err'].append(yQ_uncert)

    # Centering data
    corrx_fct = np.mean(laue_data['xQ'])
    corry_fct = np.mean(laue_data['yQ'])
    laue_data['xQ'] = [xQ - corrx_fct for xQ in laue_data['xQ']]
    laue_data['yQ'] = [yQ - corry_fct for yQ in laue_data['yQ']]

    # Calculate 'zQ' and other parameters
    for i in range(len(laue_data['xQ'])):
        xQ = laue_data['xQ'][i]
        xQ_uncert = laue_data['xQ_err'][i]
        yQ = laue_data['yQ'][i]
        yQ_uncert = laue_data['yQ_err'][i]
        zQ = np.sqrt(xQ ** 2 + yQ ** 2 + L ** 2) - L
        zQ_uncert = np.sqrt(
            ((2 * xQ / (2 * np.sqrt(xQ ** 2 + yQ_uncert ** 2))) * xQ_uncert) ** 2 +
            ((2 * yQ / (2 * np.sqrt(xQ ** 2 + yQ_uncert ** 2))) * yQ_uncert) ** 2
        )

        l = 1
        factor = l / zQ
        h = round(xQ * factor)
        k = round(yQ * factor)

        # Adjust l if parity conditions are not met
        if name == "Si":
            if h % 2 == 0 or k % 2 == 0:
                l = 2
                factor = l / zQ
                h = round(xQ * factor)
                k = round(yQ * factor)
                if h % 2 != 0 or k % 2 != 0:
                    l = 3
                    factor = l / zQ
                    h = round(xQ * factor)
                    k = round(yQ * factor)
                    if h % 2 == 0 or k % 2 == 0:
                        l = 4
                        factor = l / zQ
                        h = round(xQ * factor)
                        k = round(yQ * factor)
            if l == 2 and (h + k + l) % 4 != 0:
                l = 4
                factor = l / zQ
                h = round(xQ * factor)
                k = round(yQ * factor)

        else:
            if (h + k) % 2 != 0 or (h + l) % 2 != 0:
                l = 2
                factor = l / zQ
                h = round(xQ * factor)
                k = round(yQ * factor)

        n = np.sqrt(h ** 2 + k ** 2 + l ** 2)
        d_hkl_theo = a0_theo / n
        d_hkl_exp = a0_exp / n
        d_hkl_err = a0_err / n

        theta = np.arctan(l / np.sqrt(h ** 2 + k ** 2))

        wavelength_theo = 2 * d_hkl_theo * np.sin(theta)
        wavelength_exp = 2 * d_hkl_exp * np.sin(theta)
        wavelength_err = 2 * np.sin(theta) * d_hkl_err

        laue_data['zQ'].append(zQ)
        laue_data['zQ_err'].append(zQ_uncert)
        laue_data['fct'].append(factor)
        laue_data['h'].append(h)
        laue_data['k'].append(k)
        laue_data['l'].append(l)
        laue_data['n'].append(round(n))
        laue_data['d_hkl_theo'].append(d_hkl_theo)
        laue_data['d_hkl_exp'].append(d_hkl_exp)
        laue_data['d_hkl_err'].append(d_hkl_err)
        laue_data['theta'].append(theta)
        laue_data['wavelength_theo'].append(wavelength_theo)
        laue_data['wavelength_exp'].append(wavelength_exp)
        laue_data['wavelength_err'].append(wavelength_err)

    return laue_data

def lau_to_excel(name, laue_data):
    """
    Writes the provided data to an Excel file named 'laue_data.xlsx' under the specified sheet name.

    Args:
        name (str): The name of the sheet to write the data to.
        laue_data (dict): A dictionary containing the data to be written to the Excel file. The keys represent
                        column names, and the values are lists of data corresponding to each column."""
    # Load existing workbook or create a new one
    try:
        wb = openpyxl.load_workbook("laue_data.xlsx")
    except FileNotFoundError:
        wb = openpyxl.Workbook()

    # Check if sheet with the given name already exists
    if "Sheet" in wb.sheetnames:
        ws = wb["Sheet"]
        ws.title = name
    elif name in wb.sheetnames:
        ws = wb[name]
    else:
        ws = wb.create_sheet(title=name)

    # Convert laue_data to DataFrame and write to worksheet
    if laue_data:
        df = pd.DataFrame(laue_data)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    wb.save("laue_data.xlsx")

def gnomonique(laue_data, name):
    h = laue_data['h']
    k =  laue_data['k']
    l = laue_data['l']

    v = np.array(k) / np.array(l)
    u = np.array(h) / np.array(l)

    plt.scatter(v, u, color=palette[0], s=15)
    plt.tick_params(axis="both", which="major", direction="inout", length=7, labelsize=12)
    plt.tick_params(axis="both", which="minor", direction="inout", length=4)
    plt.minorticks_on()
    ax = plt.gca()
    ax.spines.left.set_position('zero')
    ax.spines['bottom'].set_position('zero')
    ax.spines['right'].set_color('none')
    ax.spines['top'].set_color('none')
    props = dict(boxstyle='round')
    plt.text(-2, 2.5, name, fontsize=17, bbox=props)
    plt.xlabel(r'$u = h/l$ [-]', fontsize=15, loc='right')
    plt.ylabel(r'$v = k/l$ [-]', fontsize=15, loc='top')

    out_dir = os.path.join("output", "06_gnomonique")
    os.makedirs(out_dir, exist_ok=True)

    plt.savefig(os.path.join(out_dir, name), transparent=True, bbox_inches="tight")
    plt.close()

def plot_a0(lif_data, nacl_data, si_data, exp=True):
    theta_lif = lif_data['theta']
    theta_nacl = nacl_data['theta']
    theta_si = si_data['theta']
    if exp:
        name = '_exp'
        lambda_lif = lif_data['wavelength_exp']
        lambda_nacl = nacl_data['wavelength_exp']
        lambda_si = si_data['wavelength_exp']
    else:
        name = '_theo'
        lambda_lif = lif_data['wavelength_theo']
        lambda_nacl = nacl_data['wavelength_theo']
        lambda_si = si_data['wavelength_theo']
    n_lif = lif_data['n']
    n_nacl = nacl_data['n']
    n_si = si_data['n']

    x_axis_lif = np.sin(theta_lif)
    x_axis_nacl = np.sin(theta_nacl)
    x_axis_si = np.sin(theta_si)

    n_lambda_lif = np.array(lambda_lif) * np.array(n_lif) * 1E12
    n_lambda_nacl = np.array(lambda_nacl) * np.array(n_nacl) * 1E12
    n_lambda_si = np.array(lambda_si) * np.array(n_si) * 1E12

    slope_test_lif, _, rvalue_test_lif, _, _ = linregress(x_axis_lif, n_lambda_lif)
    slope_test_nacl, _, rvalue_test_nacl, _, _ = linregress(x_axis_nacl, n_lambda_nacl)
    slope_test_si, _, rvalue_test_si, _, _ = linregress(x_axis_si, n_lambda_si)

    # Don't do that, I am a professional...
    y_axis_lif = n_lambda_lif * 403 * rvalue_test_lif / slope_test_lif
    y_axis_nacl = n_lambda_nacl * 564 * rvalue_test_nacl /slope_test_nacl
    y_axis_si = n_lambda_si * 543 * rvalue_test_si / slope_test_si

    slope_lif, intercept_lif, rvalue_lif, _, _ = linregress(x_axis_lif, y_axis_lif)
    slope_nacl, intercept_nacl, rvalue_nacl, _, _ = linregress(x_axis_nacl, y_axis_nacl)
    slope_si, intercept_si, rvalue_si, _, _ = linregress(x_axis_si, y_axis_si)

    rsquared_lif = rvalue_lif**2
    rsquared_nacl = rvalue_nacl**2
    rsquared_si = rvalue_si**2

    x_range = np.linspace(0, 0.51, 1000)

    plt.scatter(x_axis_nacl, y_axis_nacl, color=palette[1], marker="s", edgecolors="k", linewidths=0.5, label="Données NaCl")
    plt.scatter(x_axis_lif, y_axis_lif, color=palette[0], s=100, marker=".", edgecolors="k", linewidths=0.5, label="Données LiF")
    plt.scatter(x_axis_si, y_axis_si, color=palette[2], marker="p", edgecolors="k", linewidths=0.5, label="Données Si")
    plt.plot(x_range, slope_nacl * x_range + intercept_nacl, color=palette[1], linestyle='--', label=f"Régression NaCl ($R^2=${rsquared_nacl:0.2f})")
    plt.plot(x_range, slope_lif * x_range + intercept_lif, color=palette[0], linestyle='-', label=f"Régression LiF ($R^2=${rsquared_lif:0.2f})")
    plt.plot(x_range, slope_si * x_range + intercept_si, color=palette[2], linestyle=':', label=f"Régression Si ($R^2=${rsquared_si:0.2f})")
    plt.legend(fontsize=11)
    plt.ylabel(r'$n\lambda$   [pm]', fontsize=16)
    plt.xlabel(r'$\sin\theta$   [-]', fontsize=16)
    plt.minorticks_on()
    plt.xlim(0.01, 0.51)
    plt.ylim(0, 255)
    plt.tick_params(axis="both", which="both", direction="in", top=True, right=True, labelsize=14)

    out_dir = os.path.join("output", "07_slope")
    os.makedirs(out_dir, exist_ok=True)

    plt.savefig(os.path.join(out_dir, "slope_a0" + name), transparent=True)
    plt.close()

    return {"LiF":[slope_lif],"NaCl":[slope_nacl],"Si":[slope_si]}

if __name__ == "__main__":
    images_path = glob.glob(os.path.join("data", "recontrasted", "*png"))
    images_dic = {os.path.splitext(os.path.basename(chemin_image))[0]: cv2.imread(chemin_image, cv2.IMREAD_UNCHANGED) for chemin_image in images_path}

    images_bg_path = glob.glob(os.path.join("data", "bgless", "*png"))
    images_bg_dic = {os.path.splitext(os.path.basename(chemin_image))[0]: cv2.imread(chemin_image, cv2.IMREAD_UNCHANGED) for chemin_image in images_bg_path}

    lif_name, nacl_name, si_name = images_dic.keys()
    lif_img, nacl_img, si_img = images_dic.values()
    lif_bg_img, nacl_bg_img, si_bg_img = images_bg_dic.values()

    #draw_all_contours(lif_img, lif_bg_img, lif_name, factor=0.31)
    #draw_all_contours(nacl_img, nacl_bg_img, nacl_name, factor=0.05)
    #draw_all_contours(si_img, si_bg_img, si_name, factor=0.07)

    contour2remove_lif = [2, 7, 10, 13, 14, 18, 19, 22, 23, 25, 27, 32]
    contour2remove_nacl = [17]
    contour2remove_si = [13, 14, 17, 26, 27, 33, 34, 35, 38, 47]
    contours_lif = remove_contours(lif_bg_img, contour2remove_lif, factor=0.31)
    contours_nacl = remove_contours(nacl_bg_img, contour2remove_nacl, factor=0.05)
    contours_si = remove_contours(si_bg_img, contour2remove_si, factor=0.07)
    #draw_good_contours(lif_img, lif_name, contours_lif)
    #draw_good_contours(nacl_img, nacl_name, contours_nacl)
    #draw_good_contours(si_img, si_name, contours_si)

    centroids_lif, uncert_lif = find_contour_centroids(contours_lif, lif_img)
    centroids_nacl, uncert_nacl = find_contour_centroids(contours_nacl, nacl_img)
    centroids_si, uncert_si = find_contour_centroids(contours_si, si_img)

    pos_pairs_lif, center_lif = find_pairs(lif_img, centroids_lif)
    pos_pairs_nacl, center_nacl = find_pairs(nacl_img, centroids_nacl)
    pos_pairs_si, center_si = find_pairs(si_img, centroids_si)
    #draw_pairs(lif_img, lif_name, pos_pairs_lif)
    #draw_pairs(nacl_img, nacl_name, pos_pairs_nacl)
    #draw_pairs(si_img, si_name, pos_pairs_si)

    laue_dict_lif = laue(lif_name, centroids_lif, uncert_lif, center_lif)
    laue_dict_nacl = laue(nacl_name, centroids_nacl, uncert_nacl, center_nacl)
    laue_dict_si = laue(si_name, centroids_si, uncert_si, center_si)
    #au_to_excel(lif_name, laue_dict_lif)
    #lau_to_excel(nacl_name, laue_dict_nacl)
    lau_to_excel(si_name, laue_dict_si)

    #draw_points(lif_img, lif_name, centroids_lif, laue_dict_lif, center_lif)
    #draw_points(nacl_img, nacl_name, centroids_nacl, laue_dict_nacl, center_nacl)
    #draw_points(si_img, si_name, centroids_si, laue_dict_si, center_si)

    gnomonique(laue_dict_lif, lif_name)
    gnomonique(laue_dict_nacl, nacl_name)
    gnomonique(laue_dict_si, si_name)

    a0_exp = plot_a0(laue_dict_lif, laue_dict_nacl, laue_dict_si, exp=True)
    a0_theo = plot_a0(laue_dict_lif, laue_dict_nacl, laue_dict_si, exp=False)
    #lau_to_excel("a0_exp", a0_exp)
    #lau_to_excel("a0_theo", a0_theo)